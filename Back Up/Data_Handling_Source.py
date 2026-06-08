from __future__ import annotations
import numpy as np
from typing import Literal, Iterable, Mapping, Optional, Sequence, Tuple, Union, Any, Dict, List
import pandas as pd
from dataclasses import dataclass
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
import warnings
from pathlib import Path
from typing import Optional
from openpyxl.worksheet.worksheet import Worksheet
import re
from openpyxl.utils import get_column_letter
import os
import math


def normalize_column_to_column(
    file_path: str,
    sheet_name: str,
    x: str,                   # source column name
    y: str,                   # destination column name
    header_row: int = 1,
    round_digits: Optional[int] = 6,  # round results; set None to skip rounding
    create_if_missing: bool = True,
    zero_mean_atol: float = 0.0,   # <<< changed: default no false alarms
):
    """
    Normalize column `x` by dividing each numeric cell by mean(x) and write to column `y`.
    Returns the mean used.
    """

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
    ws = wb[sheet_name]

    # Map header names -> column indices
    header_cells = list(ws[header_row])
    name_to_col = {str(c.value).strip(): c.column for c in header_cells if c.value is not None}

    if x not in name_to_col:
        raise ValueError(f"Column '{x}' not found in header row {header_row}")
    src_col_idx = name_to_col[x]

    # Destination column
    if y in name_to_col:
        dst_col_idx = name_to_col[y]
    else:
        if not create_if_missing:
            raise ValueError(f"Destination column '{y}' not found and create_if_missing=False.")
        dst_col_idx = ws.max_column + 1
        ws.cell(row=header_row, column=dst_col_idx).value = y

    # Read values
    vals = []
    for r in range(header_row + 1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=src_col_idx).value
        try:
            v = float(cell_val)
        except (TypeError, ValueError):
            v = np.nan
        vals.append(v)

    arr = np.array(vals, dtype=float)
    with np.errstate(all="ignore"):
        m = np.nanmean(arr)

    # Only fail if mean is non-finite or effectively zero per chosen tolerance
    if not np.isfinite(m) or np.isclose(m, 0.0, atol=zero_mean_atol):
        raise ValueError(f"Cannot normalize: mean of '{x}' is {m} (<= atol {zero_mean_atol})")

    norm = arr / m

    # Write results
    for i, v in enumerate(norm, start=header_row + 1):
        if np.isnan(v):
            ws.cell(row=i, column=dst_col_idx).value = None
        else:
            ws.cell(row=i, column=dst_col_idx).value = round(v, round_digits) if isinstance(round_digits, int) else float(v)

    wb.save(file_path)
    return float(m)

EXCLUDE_HEADERS_DEFAULT = {"Index", "Device", "Note"}
NORMAL_SUFFIX = "_normalized"

def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    header_cells = list(ws[header_row])
    name_to_col = {}
    for c in header_cells:
        if c.value is None:
            continue
        name_to_col[str(c.value).strip()] = c.column
    return name_to_col

def _as_int(v):
    """Best-effort convert cell value to int, otherwise None."""
    if v is None:
        return None
    if isinstance(v, (int, float)) and float(v).is_integer():
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    # allow "1.0" or " 1 "
    try:
        f = float(s)
        if f.is_integer():
            return int(f)
    except ValueError:
        return None
    return None

def _iter_blocks_by_index_reset(
    ws,
    index_col: int,
    header_row: int,
    reset_value: int = 1,
) -> List[Tuple[int, int]]:
    """
    Returns list of (start_row, end_row) blocks based on Index resets.
    A new block begins when Index == reset_value (default 1).
    """
    starts: List[int] = []
    first_data = header_row + 1
    last = ws.max_row

    for r in range(first_data, last + 1):
        idx = _as_int(ws.cell(r, index_col).value)
        if idx == reset_value:
            starts.append(r)

    if not starts:
        return []

    blocks: List[Tuple[int, int]] = []
    for i, s in enumerate(starts):
        e = (starts[i + 1] - 1) if (i + 1 < len(starts)) else last
        # optional tightening: ignore trailing completely empty rows
        while e >= s and all(ws.cell(e, c).value in (None, "") for c in range(1, ws.max_column + 1)):
            e -= 1
        if e >= s:
            blocks.append((s, e))

    return blocks


def add_block_id_column(
    file_path: str | Path,
    sheet_name: str,
    *,
    header_row: int = 1,
    index_header: str = "Index",
    block_header: str = "BlockID",
    reset_value: int = 1,
    overwrite: bool = True
):
    wb = load_workbook(str(file_path))
    ws = wb[sheet_name]
    name_to_col = _build_header_map(ws, header_row)

    if index_header not in name_to_col:
        raise ValueError(f"Header '{index_header}' not found.")
    index_col = name_to_col[index_header]

    if not overwrite and block_header in name_to_col:
        raise ValueError("BlockID column already exists and overwrite=False.")

    # create BlockID column if missing
    if block_header in name_to_col:
        block_col = name_to_col[block_header]
    else:
        block_col = ws.max_column + 1
        ws.cell(header_row, block_col).value = block_header

    blocks = _iter_blocks_by_index_reset(ws, index_col=index_col, header_row=header_row, reset_value=reset_value)

    for b, (r0, r1) in enumerate(blocks, start=1):
        for r in range(r0, r1 + 1):
            ws.cell(r, block_col).value = b

    wb.save(str(file_path))



def normalize_params_blockwise(
    file_path: str | Path,
    sheet_name: str,
    *,
    header_row: int = 1,
    suffix: str = NORMAL_SUFFIX,
    exclude_headers: Optional[Iterable[str]] = None,
    round_digits: Optional[int] = 6,
    create_if_missing: bool = True,
    zero_mean_atol: float = 0.0,
) -> Dict[Tuple[int, int], Dict[str, float]]:
    """
    Block-wise normalization for a sheet that contains multiple stacked blocks.

    For each block:
      normalized = value / mean(block_values) for each numeric column.
    Means are computed per block, per parameter.

    Returns:
      {(block_start, block_end): {param_name: mean_used}}
    """
    file_path = str(file_path)
    exclude = set(exclude_headers) if exclude_headers is not None else set(EXCLUDE_HEADERS_DEFAULT)

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
    ws = wb[sheet_name]

    name_to_col = _build_header_map(ws, header_row)

    # Detect blocks
    if "Index" not in name_to_col:
        raise ValueError("Header 'Index' not found. Index-reset block detection needs Index column.")
    index_col = name_to_col["Index"]

    blocks = _iter_blocks_by_index_reset(ws, index_col=index_col, header_row=header_row, reset_value=1)

    src_params = []
    for h in name_to_col.keys():
        if h in exclude:
            continue
        if h.endswith(suffix):
            continue
        src_params.append(h)

    # Ensure destination normalized columns exist (create if missing)
    # IMPORTANT: keep mapping for ALL params used in sheet, even if some blocks have blanks
    norm_col: Dict[str, int] = {}
    for p in src_params:
        dst = f"{p}{suffix}"
        if dst in name_to_col:
            norm_col[p] = name_to_col[dst]
        else:
            if not create_if_missing:
                raise ValueError(f"Destination column '{dst}' not found and create_if_missing=False.")
            # append at end
            new_c = ws.max_column + 1
            ws.cell(header_row, new_c).value = dst
            norm_col[p] = new_c
            name_to_col[dst] = new_c

    means_by_block: Dict[Tuple[int, int], Dict[str, float]] = {}

    # Normalize each block independently
    for (r0, r1) in blocks:
        means_this_block: Dict[str, float] = {}

        for p in src_params:
            src_c = name_to_col[p]
            dst_c = norm_col[p]

            # collect numeric values in this block
            row_vals = []  # keep per-row numeric or nan for writing
            for r in range(r0, r1 + 1):
                cell_val = ws.cell(r, src_c).value
                try:
                    v = float(cell_val)
                except (TypeError, ValueError):
                    v = np.nan
                row_vals.append(v)
            arr = np.array(row_vals, dtype=float)

            # ---- NEW SAFE CHECK (put HERE) ----
            finite_vals = arr[np.isfinite(arr)]
            if finite_vals.size == 0:
                # Column exists but has no numeric data in this block
                for r in range(r0, r1 + 1):
                    ws.cell(r, dst_c).value = None
                continue

            m = finite_vals.mean()

            if np.isclose(m, 0.0, atol=zero_mean_atol):
                for r in range(r0, r1 + 1):
                    ws.cell(r, dst_c).value = None
                continue

            means_this_block[p] = float(m)
            norm = arr / m
            # norm = np.array(row_vals, dtype=float) / m


            # write normalized values for this block only
            for idx, v in enumerate(norm, start=r0):
                if np.isnan(v):
                    ws.cell(idx, dst_c).value = None
                else:
                    ws.cell(idx, dst_c).value = round(v, round_digits) if isinstance(round_digits, int) else float(v)

        means_by_block[(r0, r1)] = means_this_block

    wb.save(file_path)
    return means_by_block







@dataclass
class VariableVector:
    x_name: str               # e.g. "x1"
    input_type: str           # e.g. "Device" / "Parameter"
    value_name: str           # e.g. "Q1_value"
    name: str                 # e.g. "x1_Q1_value" (or you can include input_type)
    values: List[Any]         # extracted values
    size: int                 # len(values)

@dataclass
class CaptureResult:
    variables_with_data: List[str]                  # e.g. ["x1", "x2", "x3"]
    vectors: Dict[str, VariableVector]              # vector_name -> VariableVector
    sizes: Dict[str, int]                           # vector_name -> size
    input_types: Dict[str, str]                     # vector_name -> input_type
    by_input_type: Dict[str, List[str]]             # input_type -> [vector_name,...]
    total_variables_with_data: int                  # len(variables_with_data)

def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False
def capture_excel_variables(
    xlsx_path: str,
    sheet_name: Optional[str] = None,
    *,
    order_col: str = "A",
    x_start_col: str = "B",        # x1
    x_end_col: str = "M",          # x12
    input_type_row: int = 2,     # NEW: row containing Input_Type values
    value_name_row: int = 3,       # row containing Q1_value / SwitchingFrequency / ...
    first_data_row: int = 4,
    include_blanks: bool = False,  # False => drop empty cells from vectors
) -> CaptureResult:
    """
    1) Detect which x-columns contain any data in the data region.
    2) For each x-column with data, build a vector named: "<x#>_<Value_Name>".
       Example: x1_Q1_value
    3) Return sizes and a summary structure.
    """
    warnings.filterwarnings(
        "ignore",
        message="Data Validation extension is not supported*"
    )
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    order_c = column_index_from_string(order_col)
    x_start = column_index_from_string(x_start_col)
    x_end = column_index_from_string(x_end_col)

    # Determine last row by scanning down the Order column (stop after a run of blanks).
    # This is robust if the sheet has formatted empty rows.
    last_row = ws.max_row
    # Optional tightening: find last non-empty order entry
    for r in range(ws.max_row, first_data_row - 1, -1):
        if not _is_empty(ws.cell(r, order_c).value):
            last_row = r
            break

    variables_with_data: List[str] = []
    vectors: Dict[str, VariableVector] = {}
    sizes: Dict[str, int] = {}
    input_types: Dict[str, str] = {}
    by_input_type: Dict[str, List[str]] = {}

    for col in range(x_start, x_end + 1):
        x_header = ws.cell(1, col).value  # row 1: "x1", "x2", ...
        if _is_empty(x_header):
            continue
        #x_name = str(x_header).strip()

        # NEW: read input_type from row 2
        input_type = ws.cell(input_type_row, col).value
        if _is_empty(input_type):
            input_type = "Unknown"
        input_type = str(input_type).strip()

        value_name = ws.cell(value_name_row, col).value  # row 3: "Q1_value", etc.
        if _is_empty(value_name):
            # fallback if missing
            value_name = "Value"
        value_name = str(value_name).strip()

        # Collect values down the data region
        col_values: List[Any] = []
        has_any_data = False

        for r in range(first_data_row, last_row + 1):
            v = ws.cell(r, col).value
            if _is_empty(v):
                if include_blanks:
                    col_values.append(None)
                continue
            has_any_data = True
            col_values.append(v)

        if has_any_data:
            x_name = str(x_header).strip()
            variables_with_data.append(x_name)

            vector_name = f"{x_name}_{str(value_name).strip()}"
            # vec = VariableVector(name=vector_name, values=col_values, size=len(col_values))
            vec = VariableVector(
                x_name=x_name,
                input_type=input_type,
                value_name=value_name,
                name=vector_name,
                values=col_values,
                size=len(col_values),
            )
            vectors[vector_name] = vec
            sizes[vector_name] = vec.size
            input_types[vector_name] = input_type
            by_input_type.setdefault(input_type, []).append(vector_name)

    return CaptureResult(
        variables_with_data=variables_with_data,
        vectors=vectors,
        sizes=sizes,
        input_types=input_types,
        by_input_type=by_input_type,
        total_variables_with_data=len(variables_with_data),
    )

def build_variable_report(result: CaptureResult) -> str:
    """
    Human-readable report:
    1) How many variables with data?
    2) Size of each variable vector
    """
    lines: List[str] = []
    lines.append(f"Variables with data: {result.total_variables_with_data}")
    lines.append(f"Variables: {', '.join(result.variables_with_data) if result.variables_with_data else '(none)'}")
    lines.append("")
    lines.append("Vector sizes:")
    for name, size in result.sizes.items():
        lines.append(f" - {name}: {size}")
    return "\n".join(lines)


#Reading data from Chara file into the UserProvidedDatafile


# -----------------------------
# Configuration / Metadata
# -----------------------------

# Map: parameter name -> sheet name inside "chara" workbook
# Extend this mapping as your library grows.
DEFAULT_PARAM_TO_SHEET: Dict[str, str] = {
    "Ciss": "Capacitance",
    "Coss": "Capacitance",
    "Crss": "Capacitance",
    "Rds(on)": "Rds(on)",
    "Rg": "Rg",
    "VTH": "VTH",
    "body diode": "body diode",
    "Transfer": "Transfer",
    "Output": "Output",
    "Ciss_curve": "Ciss",  # if you have separate tabs named like "Ciss"
    "Coss_curve": "Coss",
    "Crss_curve": "Crss",
    "Eon": "Eon",
    "Eoff": "Eoff",
    "ID_25": "ID_25",
    "VF":"VF",
    "Rd":"Rd",
    "IF":"IF",
    "Qc":"Qc",
    "Ae":"Ae",
    "le":"le",
    "k1":"k1",
    "k2":"k2",
    "anpha":"anpha",
    "beta":"beta",
    "b":"b",
    "c":"c",
}

block_specs = [
    {"block_id": 1, "z_col": 0,
     "cols": ["Rds(on)_normalized", "VTH_normalized", "ID_25_normalized", "Eon_normalized", "Eoff_normalized"]},
    {"block_id": 2, "z_col": 1, "cols": ["VF_normalized","Rd_normalized", "IF_normalized", "Qc_normalized"]},
    {"block_id": 3, "z_col": 2, "cols": ["Ae_normalized", "le_normalized", "k1_normalized","k2_normalized","anpha_normalized", "beta_normalized","b_normalized","c_normalized"]},
]

@dataclass
class ExtractedParam:
    device: str
    param: str
    value: Optional[float]     # numeric if possible; None if not found/parseable
    unit: Optional[str]        # if detected; else None
    source_sheet: Optional[str]
    source_cell: Optional[str] # Excel address like "B2" when found
    note: Optional[str]        # reason when missing / ambiguous


@dataclass
class DeviceResult:
    device: str
    ok: bool
    chara_path: Optional[str]
    params: Dict[str, ExtractedParam]  # param -> ExtractedParam
    note: Optional[str]                # device-level note (missing folder/file etc.)


# -----------------------------
# Helpers
# -----------------------------

def _normalize(s: Any) -> str:
    return str(s).strip() if s is not None else ""


def _find_chara_file(device_folder: Path) -> Optional[Path]:
    """
    Find an Excel file whose name contains 'chara' (case-insensitive) in the device folder.
    Preference: .xlsx first, then .xlsm.
    """
    candidates = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(device_folder.glob(ext))
    candidates = [p for p in candidates if "chara" in p.name.lower()]
    if not candidates:
        return None
    # If multiple, pick the newest by modified time
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]

def _coerce_float(x: Any) -> Optional[float]:
    """
    Attempt to convert a value to float.
    Handles numbers directly and strings like '1203', '1203 pF', '1,203', etc.
    Returns None if not parseable.
    """
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = _normalize(x)
    if not s:
        return None

    # Remove commas and common unit strings; keep sign/dot/exponent digits
    s2 = s.replace(",", " ")
    # Split by whitespace and try first numeric-like token
    tokens = s2.split()
    for t in tokens:
        try:
            return float(t)
        except ValueError:
            continue

    # Try stripping non-numeric characters more aggressively
    import re
    m = re.search(r"[-+]?\d+(\.\d+)?([eE][-+]?\d+)?", s2)
    if m:
        try:
            return float(m.group(0))
        except ValueError:
            return None
    return None


def _guess_unit_nearby(ws: Worksheet, row: int, col: int) -> Optional[str]:
    """
    Heuristic: look in the same row near the value cell (or a few columns right)
    for a short string that looks like a unit, e.g. 'pF', 'nF', 'mΩ', 'V', 'A', 'µJ'.
    """
    unit_candidates = []
    for dc in range(1, 5):
        v = ws.cell(row=row, column=col + dc).value
        s = _normalize(v)
        if 0 < len(s) <= 6:  # short strings
            unit_candidates.append(s)

    # Pick the first that looks unit-like
    for u in unit_candidates:
        # very lightweight check
        if any(sym in u for sym in ["pF", "nF", "uF", "µF", "mΩ", "Ohm", "Ω", "V", "A", "W", "J", "µJ", "mJ"]):
            return u
    return None


def _find_param_value_in_sheet(
    ws: Worksheet,
    param: str,
    *,
    max_scan_rows: int = 200,
    max_scan_cols: int = 40,
) -> Tuple[Optional[float], Optional[str], Optional[str], Optional[str]]:
    """
    Generic extractor:
    - Scan a bounded area (top-left) for a cell whose text equals param (case-insensitive).
    - If found, attempt to read the numeric value from:
        (same row, next columns) or (next row, same column), whichever yields a number first.
    Returns: (value, unit, cell_addr_of_value, note)
    """
    target = param.strip().lower()

    # Bound scan to avoid reading huge sheets
    r_max = min(ws.max_row or 0, max_scan_rows)
    c_max = min(ws.max_column or 0, max_scan_cols)

    for r in range(1, r_max + 1):
        for c in range(1, c_max + 1):
            cell_val = ws.cell(r, c).value
            if _normalize(cell_val).lower() == target:
                # Candidate location for value: rightwards then downward
                # 1) same row, next 1..6 columns
                for dc in range(1, 7):
                    v = ws.cell(r, c + dc).value
                    fv = _coerce_float(v)
                    if fv is not None:
                        unit = _guess_unit_nearby(ws, r, c + dc)
                        addr = ws.cell(r, c + dc).coordinate
                        return fv, unit, addr, None

                # 2) next rows, same column / next columns (sometimes tables are vertical)
                for dr in range(1, 4):
                    v = ws.cell(r + dr, c).value
                    fv = _coerce_float(v)
                    if fv is not None:
                        unit = _guess_unit_nearby(ws, r + dr, c)
                        addr = ws.cell(r + dr, c).coordinate
                        return fv, unit, addr, None

                return None, None, None, f"Found '{param}' label but no numeric value adjacent."

    return None, None, None, f"Parameter '{param}' not found in scanned area."


# -----------------------------
# Main function
# -----------------------------
def clear_destination_sheet(
    dest_xlsx_path: str | Path,
    dest_sheet_name: str,
    create_if_missing: bool = True,
) -> None:
    dest_path = Path(dest_xlsx_path)

    # Open or create workbook
    if dest_path.exists():
        wb = load_workbook(dest_path)
    else:
        if not create_if_missing:
            raise FileNotFoundError(dest_path)
        wb = Workbook()

    # Get or create sheet
    if dest_sheet_name in wb.sheetnames:
        ws = wb[dest_sheet_name]
        # Clear everything
        ws.delete_rows(1, ws.max_row)
        ws.delete_cols(1, ws.max_column)
    else:
        wb.create_sheet(dest_sheet_name)

    wb.save(dest_path)


def extract_params_from_chara_library(
    *,
    available_devices: Iterable[str],
    requested_devices: Optional[Iterable[str]] = None,
    requested_params: Iterable[str],
    library_root_dir: str | Path,
    dest_xlsx_path: str | Path,
    dest_sheet_name: str,
    param_to_sheet: Optional[Dict[str, str]] = None,
    create_dest_if_missing: bool = True,
) -> List[DeviceResult]:
    """
    1) Check device part number in available list.
    2) For each device, go to folder <library_root_dir>/<device>/.
    3) Find an Excel file containing "chara" in the filename.
    4) For each requested parameter, open the mapped sheet and extract value.
    5) Write results into destination Excel sheet (device rows, parameter columns).
    6) Return detailed results for logging / debugging.

    Output format in destination sheet:
      Col A: Device
      Col B: requested parameters (in the order provided)
      Last Col: Note (device-level), optional
    """
    lib_root = Path(library_root_dir)
    dest_path = Path(dest_xlsx_path)
    param_to_sheet = param_to_sheet or DEFAULT_PARAM_TO_SHEET

    available_set = {d.strip() for d in available_devices if str(d).strip()}
    if requested_devices is None:
        devices = [d.strip() for d in available_devices if str(d).strip()]
    else:
        devices = [d.strip() for d in requested_devices if str(d).strip()]

    params = [p.strip() for p in requested_params if str(p).strip()]

    results: List[DeviceResult] = []

    # Prepare destination workbook
    if dest_path.exists():
        try:
            with open(dest_path, "a"):
                pass
        except PermissionError:
            raise RuntimeError(
                f"Destination Excel file is open in Excel. "
                f"Please close it before running extraction:\n{dest_path}"
            )
        dest_wb = load_workbook(dest_path)
    else:
        if not create_dest_if_missing:
            raise FileNotFoundError(f"Destination workbook not found: {dest_path}")
        dest_wb = Workbook()

    if dest_sheet_name in dest_wb.sheetnames:
        ws_out = dest_wb[dest_sheet_name]
    else:
        ws_out = dest_wb.create_sheet(dest_sheet_name)

    # Internal helper 00
    def _get_or_create_header_col(ws, header: str) -> int:
        # find existing header
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None and str(v).strip() == header:
                return c
        # not found -> append new column
        new_c = ws.max_column + 1
        ws.cell(1, new_c, header)
        return new_c

    #Internal helper 01
    def _next_empty_row(ws) -> int:
        # returns the first row where column 2 (Device column) is empty
        r = ws.max_row
        if r < 2:
            return 2
        for rr in range(r, 1, -1):
            if ws.cell(rr, 2).value not in (None, ""):
                return rr + 1
        return 2

    # Internal helper 02
    def _ensure_header_and_get_colmap(ws, params: list[str]) -> Dict[str, int]:
        # 1) If sheet empty, create base headers
        if ws.cell(1, 1).value in (None, ""):
            ws.cell(1, 1, "Index")
            ws.cell(1, 2, "Device")
            ws.cell(1, 3, "Note")

        # helper: find column by header name
        def find_col(name: str) -> int | None:
            for c in range(1, ws.max_column + 1):
                v = ws.cell(1, c).value
                if v is not None and str(v).strip() == name:
                    return c
            return None

        # 2) Ensure Index + Device exist
        if find_col("Index") is None:
            ws.insert_cols(1)
            ws.cell(1, 1, "Index")
        if find_col("Device") is None:
            ws.insert_cols(2)
            ws.cell(1, 2, "Device")

        # 3) Ensure Note exists
        note_col = find_col("Note")
        if note_col is None:
            ws.cell(1, ws.max_column + 1, "Note")
            note_col = ws.max_column

        # 4) Ensure all params exist; if missing -> insert BEFORE Note
        for p in params:
            if find_col(p) is None:
                note_col = find_col("Note")  # re-find each time
                ws.insert_cols(note_col)
                ws.cell(1, note_col, p)

        # 5) FORCE Note to be the LAST column (move it if needed)
        note_col = find_col("Note")
        last_col = ws.max_column
        if note_col is not None and note_col != last_col:
            # copy Note column to end then delete old Note column
            new_col = last_col + 1
            ws.cell(1, new_col, "Note")
            for r in range(2, ws.max_row + 1):
                ws.cell(r, new_col, ws.cell(r, note_col).value)
            ws.delete_cols(note_col)

        # 6) Build final colmap
        colmap: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is None:
                continue
            colmap[str(v).strip()] = c

        return colmap

    # Internal helper 03


    #End of internal help 01-02-03-04
    colmap = _ensure_header_and_get_colmap(ws_out, params)
    index_col = colmap["Index"]
    device_col = colmap["Device"]
    note_col = colmap["Note"]
    out_row = _next_empty_row(ws_out)
    block_index = 1  # or 0 if you want 0-based index

    # Internal helper 05
    def _write_output_row(device: str, values_by_param: Dict[str, Any], note: str) -> None:
        nonlocal out_row, block_index

        ws_out.cell(out_row, index_col, block_index)
        ws_out.cell(out_row, device_col, device)

        # write each requested param into its correct column (from colmap)
        for p in params:
            ws_out.cell(out_row, colmap[p], values_by_param.get(p))

        ws_out.cell(out_row, note_col, note)

        out_row += 1
        block_index += 1

    # End of Internal helper 05

    for dev in devices:
        dev_params: Dict[str, ExtractedParam] = {}
        device_note: Optional[str] = None
        chara_path: Optional[str] = None

        # (1) availability check
        if dev not in available_set:
            device_note = "Device not in available_devices list."
            _write_output_row(
                device=dev,
                values_by_param={},  # no data
                note=device_note
            )
            # still write a row for traceability
            for p in params:
                dev_params[p] = ExtractedParam(
                    device=dev, param=p, value=None, unit=None,
                    source_sheet=None, source_cell=None, note=device_note
                )
            results.append(DeviceResult(device=dev, ok=False, chara_path=None, params=dev_params, note=device_note))
            continue

        # (2) device folder
        dev_folder = lib_root / dev
        if not dev_folder.exists() or not dev_folder.is_dir():
            device_note = f"Device folder not found: {dev_folder}"
            _write_output_row(
                device=dev,
                values_by_param={},
                note=device_note
            )
            for p in params:
                dev_params[p] = ExtractedParam(
                    device=dev, param=p, value=None, unit=None,
                    source_sheet=None, source_cell=None, note=device_note
                )
            results.append(DeviceResult(device=dev, ok=False, chara_path=None, params=dev_params, note=device_note))
            continue

        # (3) find chara excel
        chara_file = _find_chara_file(dev_folder)
        if chara_file is None:
            device_note = "No Excel file containing 'chara' found in device folder."
            _write_output_row(
                device=dev,
                values_by_param={},
                note=device_note
            )
            for p in params:
                dev_params[p] = ExtractedParam(
                    device=dev, param=p, value=None, unit=None,
                    source_sheet=None, source_cell=None, note=device_note
                )
            results.append(DeviceResult(device=dev, ok=False, chara_path=None, params=dev_params, note=device_note))
            continue

        chara_path = str(chara_file)

        # Open chara workbook once per device
        wb = load_workbook(chara_file, data_only=True)

        # (4)(5) extract each param
        ok_device = True
        for p in params:
            sheet = param_to_sheet.get(p)
            if sheet is None:
                ok_device = False
                dev_params[p] = ExtractedParam(
                    device=dev, param=p, value=None, unit=None,
                    source_sheet=None, source_cell=None,
                    note=f"No sheet mapping for parameter '{p}'."
                )
                continue

            if sheet not in wb.sheetnames:
                ok_device = False
                dev_params[p] = ExtractedParam(
                    device=dev, param=p, value=None, unit=None,
                    source_sheet=sheet, source_cell=None,
                    note=f"Sheet '{sheet}' not found in chara workbook."
                )
                continue

            ws = wb[sheet]
            val, unit, addr, note = _find_param_value_in_sheet(ws, p)
            if val is None:
                ok_device = False
            dev_params[p] = ExtractedParam(
                device=dev, param=p, value=val, unit=unit,
                source_sheet=sheet, source_cell=addr, note=note
            )

        # Device-level note
        if ok_device:
            device_note = "OK"
        else:
            device_note = "One or more parameters missing / not parsed."

        results.append(DeviceResult(device=dev, ok=ok_device, chara_path=chara_path, params=dev_params, note=device_note))

        # (6) write to destination sheet row
        # write to destination sheet row
        row_values = {p: dev_params[p].value for p in params}
        _write_output_row(dev, row_values, device_note)

    # Save destination workbook
    dest_wb.save(dest_path)

    return results


def _is_selected_y(v: Any) -> bool:
    """Return True if the cell indicates selection (Y/y)."""
    if v is None:
        return False
    return str(v).strip().lower() == "y"


def _parse_weight(cell_value) -> Optional[float]:
    """
    Interprets the 'selection cell' in your device sheet.

    Rules:
      - 'Y'/'y' => selected with default weight 1.0
      - numeric (int/float or numeric string) => selected if > 0, weight = that value
      - blank / None / 0 / non-numeric text => not selected (return None)
    """
    if cell_value is None:
        return None

    # Strings: could be 'Y' or '0.4'
    if isinstance(cell_value, str):
        s = cell_value.strip()
        if s == "":
            return None
        if s.lower() == "y":
            return 1.0
        # Try numeric string
        try:
            w = float(s)
            return w if w > 0 else None
        except ValueError:
            return None

    # Numeric types
    if isinstance(cell_value, (int, float)):
        return float(cell_value) if float(cell_value) > 0 else None

    # Anything else => not selected
    return None


def capture_requested_params_and_weights_from_device_sheet(
    root_xlsx_path: str,
    *,
    device_sheet_name: str = "device",
    q_header_row: int = 2,          # row containing Q1..Q8
    q_start_col: int = 3,           # column C = 3
    param_name_col: int = 2,        # column B = 2
    first_param_row: int = 3,       # row 3 begins parameters
) -> Dict[str, Dict[str, List[Union[str, float]]]]:
    """
    Parse your provided 'device' sheet structure.

    Returns:
      {
        "Q1": {"params": [...], "weights": [...]},
        "Q2": {"params": [...], "weights": [...]},
        ...
      }

    Notes:
      - Sheet A behavior: 'Y' => weight 1.0
      - Sheet B behavior: numeric entry => that numeric weight (must be > 0)
    """
    wb = load_workbook(root_xlsx_path, data_only=True)
    if device_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{device_sheet_name}' not found in workbook: {root_xlsx_path}")

    ws = wb[device_sheet_name]

    # 1) Read Q headers from row 2, starting col C, until blank
    q_cols: Dict[str, int] = {}
    col = q_start_col
    while True:
        q = ws.cell(q_header_row, col).value
        if q is None or str(q).strip() == "":
            break
        q_name = str(q).strip()  # "Q1", "Q2", ...
        q_cols[q_name] = col
        col += 1

    if not q_cols:
        raise ValueError("No Q headers found (expected Q1.. in row 2 starting at column C).")

    # 2) For each parameter row, collect selected params + weights per Q
    q_to: Dict[str, Dict[str, List[Union[str, float]]]] = {
        q: {"params": [], "weights": []} for q in q_cols.keys()
    }

    for row in range(first_param_row, ws.max_row + 1):
        param = ws.cell(row, param_name_col).value
        if param is None or str(param).strip() == "":
            continue  # or break if your sheet ends here

        param_name = str(param).strip()

        for q_name, q_col in q_cols.items():
            raw = ws.cell(row, q_col).value
            w = _parse_weight(raw)
            if w is not None:
                q_to[q_name]["params"].append(param_name)
                q_to[q_name]["weights"].append(w)

    return q_to


# def capture_requested_params_from_device_sheet(
#     root_xlsx_path: str,
#     *,
#     device_sheet_name: str = "device",
#     q_header_row: int = 2,          # row containing Q1..Q8
#     q_start_col: int = 3,           # column C = 3
#     param_name_col: int = 2,        # column B = 2
#     first_param_row: int = 3,       # row 3 begins parameters
# ) -> Dict[str, List[str]]:
#     """
#     Parse your provided 'device' sheet structure.
#
#     Returns:
#         {
#           "Q1": ["Rds(on)", "Coss", ...],
#           "Q2": ["Rds(on)", "Coss", "Crss", "Rg", ...],
#           ...
#         }
#     """
#     wb = load_workbook(root_xlsx_path, data_only=True)
#     if device_sheet_name not in wb.sheetnames:
#         raise ValueError(f"Sheet '{device_sheet_name}' not found in workbook: {root_xlsx_path}")
#
#     ws = wb[device_sheet_name]
#
#     # 1) Read Q headers from row 2, starting col C, until blank
#     q_cols: Dict[str, int] = {}
#     col = q_start_col
#     while True:
#         q = ws.cell(q_header_row, col).value
#         if q is None or str(q).strip() == "":
#             break
#         q_name = str(q).strip()  # "Q1", "Q2", ...
#         q_cols[q_name] = col
#         col += 1
#
#     if not q_cols:
#         raise ValueError("No Q headers found (expected Q1.. in row 2 starting at column C).")
#
#     # 2) For each parameter row, check Y under each Q column
#     q_to_params: Dict[str, List[str]] = {q: [] for q in q_cols.keys()}
#
#     for row in range(first_param_row, ws.max_row + 1):
#         param = ws.cell(row, param_name_col).value
#         if param is None or str(param).strip() == "":
#             # you can choose to stop at first blank param; for now we just skip
#             continue
#
#         param_name = str(param).strip()
#
#         for q_name, q_col in q_cols.items():
#             flag = ws.cell(row, q_col).value
#             if _is_selected_y(flag):
#                 q_to_params[q_name].append(param_name)
#
#     return q_to_params


def _value_name_to_q(value_name: str) -> str:
    """
    Convert your Input sheet Value_Name into Q key used in 'device' sheet.
    Examples:
      "Q1_value" -> "Q1"
      "Q2_value" -> "Q2"
      "Q3"       -> "Q3"
    """
    s = str(value_name).strip()
    if s.endswith("_value"):
        s = s[:-len("_value")]
    return s

def build_normalized_features_by_q(
    q_to_params: Dict[str, List[str]],
    *,
    suffix: str = "_normalized",
    drop_empty: bool = True,
) -> Dict[str, List[str]]:
    """
    Convert {Q: [param, ...]} into {Q: [param_normalized, ...]}.

    Example:
        {"Q1": ["Rds(on)", "Coss"]}
        → {"Q1": ["Rds(on)_normalized", "Coss_normalized"]}
    """
    features_by_q: Dict[str, List[str]] = {}

    for q, params in q_to_params.items():
        if not params:
            if not drop_empty:
                features_by_q[q] = []
            continue

        features_by_q[q] = [f"{p}{suffix}" for p in params]

    return features_by_q

def build_requested_params_per_device_vector(
    res,  # your CaptureResult from capture_excel_variables(...)
    q_to_params: Dict[str, List[str]],
    *,
    device_input_type: str = "Device",
) -> Dict[str, List[str]]:
    """
    Returns:
      { "x1_Q1_value": [...], "x4_Q2_value": [...], ... }
    Only for vectors where input_type == "Device".
    """
    out: Dict[str, List[str]] = {}

    for vec_name, vec in res.vectors.items():
        if str(vec.input_type).strip().lower() != device_input_type.lower():
            continue

        # vec.value_name should be the Value_Name in row 3, e.g. "Q1_value"
        q_key = _value_name_to_q(vec.value_name)  # => "Q1"

        out[vec_name] = q_to_params.get(q_key, [])

    return out


class ConfigError(ValueError):
    """Raised when the Excel configuration is invalid."""

@dataclass
class ExcelConfig:
    optimizing_method: str
    init_condition_setting: str              # "Manual" or "Automated"
    initial_design_options: str              # e.g., "Index"
    xinit_indices: Dict[str, int]            # {"Xinit1": 1, "Xinit2": 1920, ...}

    dictionary_setting: str                  # "Manual" or "Automated"
    dictionary_q_values: Dict[str, List[float]]  # {"Q1_value": [1,3,5,...], ...}


def _norm_text(x) -> str:
    return str(x).strip() if x is not None else ""


def _find_key_rows(ws, col: int = 1, max_row: int = 200) -> Dict[str, int]:
    """
    Build a mapping: key_text -> row_index (1-based).
    Looks down a single column (default col A).
    """
    key_to_row = {}
    for r in range(1, max_row + 1):
        key = _norm_text(ws.cell(r, col).value)
        if key:
            key_to_row[key] = r
    return key_to_row

def _get_value_next_to_key(ws, key_to_row: Dict[str, int], key: str, value_col: int = 2) -> str:
    if key not in key_to_row:
        raise ConfigError(f"Missing required key '{key}' in column A.")
    r = key_to_row[key]
    return _norm_text(ws.cell(r, value_col).value)


def _parse_manual_or_auto(label: str, raw: str) -> str:
    v = raw.strip().lower()
    if v in ("manual", "manually"):
        return "Manual"
    if v in ("auto", "automated", "automatic"):
        return "Automated"
    raise ConfigError(f"{label} must be 'Manual' or 'Automated' (got '{raw}').")


def _parse_xinit_indices(ws, key_to_row: Dict[str, int], value_col: int = 2) -> Dict[str, int]:
    """
    Finds consecutive keys like Xinit1, Xinit2, ... in column A.
    Returns only those that exist, sorted by index number.
    """
    xinit_pattern = re.compile(r"^Xinit(\d+)$", re.IGNORECASE)
    found: List[Tuple[int, str, int]] = []

    for key, r in key_to_row.items():
        m = xinit_pattern.match(key)
        if not m:
            continue
        idx_num = int(m.group(1))
        raw = ws.cell(r, value_col).value
        if raw is None or _norm_text(raw) == "":
            # Keep as missing for validation later
            found.append((idx_num, key, None))  # type: ignore
        else:
            try:
                found.append((idx_num, key, int(raw)))
            except Exception:
                raise ConfigError(f"'{key}' must be an integer index (got '{raw}').")

    found.sort(key=lambda x: x[0])

    # Return mapping; if some are None we still include them to trigger validation upstream if needed
    out: Dict[str, int] = {}
    for _, key, val in found:
        if val is not None:
            out[key] = val
    return out

def _read_dictionary_table(
    ws,
    header_row: int = 15,
    start_col: int = 1,
    end_col: int = 8,
    data_start_row: int = 16,
    max_rows: int = 500,
) -> Dict[str, List[float]]:
    """
    Reads Q columns. A column is considered "used" if it has at least one numeric value.
    Stops reading each column when it hits a long run of blanks (simple heuristic).
    """
    headers = []
    for c in range(start_col, end_col + 1):
        headers.append(_norm_text(ws.cell(header_row, c).value))

    if not any(headers):
        raise ConfigError(f"Dictionary header row {header_row} appears empty.")

    q_map: Dict[str, List[float]] = {}
    for c, h in zip(range(start_col, end_col + 1), headers):
        if not h:
            continue

        values: List[float] = []
        blank_streak = 0

        for r in range(data_start_row, min(data_start_row + max_rows, ws.max_row + 1)):
            v = ws.cell(r, c).value
            if v is None or _norm_text(v) == "":
                blank_streak += 1
                # if there are 5 consecutive blanks, assume end of that column’s list
                if blank_streak >= 5:
                    break
                continue

            blank_streak = 0
            # accept numeric types or numeric strings
            try:
                values.append(float(v))
            except Exception:
                raise ConfigError(f"Non-numeric value in dictionary column '{h}' at {r},{c}: '{v}'")

        # mark column as "used" only if it has numeric entries
        if values:
            q_map[h] = values

    return q_map


def parse_excel_config(
    file_path: str,
    sheet_name: str,
    *,
    max_scan_rows: int = 200,
) -> ExcelConfig:
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ConfigError(f"Sheet '{sheet_name}' not found in '{file_path}'")
    ws = wb[sheet_name]

    key_to_row = _find_key_rows(ws, col=1, max_row=max_scan_rows)

    optimizing_method = _get_value_next_to_key(ws, key_to_row, "OptimizingMethod", value_col=2)

    init_setting_raw = _get_value_next_to_key(ws, key_to_row, "Initial Condition Setting", value_col=2)
    init_condition_setting = _parse_manual_or_auto("Initial Condition Setting", init_setting_raw)

    initial_design_options = _get_value_next_to_key(ws, key_to_row, "Initial Design Options", value_col=2)

    # Xinit indices
    xinit_indices = _parse_xinit_indices(ws, key_to_row, value_col=2)

    # Validation rule: if manual, user must provide at least one Xinit index (and typically all required ones)
    if init_condition_setting == "Manual":
        if not xinit_indices:
            raise ConfigError(
                "Initial Condition Setting is Manual, but no Xinit indices were provided. "
                "Please fill B4, B5, ... for Xinit1, Xinit2, ..."
            )

    # Dictionary setting block
    dict_setting_raw = _get_value_next_to_key(ws, key_to_row, "Dictionary Setting", value_col=2)
    dictionary_setting = _parse_manual_or_auto("Dictionary Setting", dict_setting_raw)

    dictionary_q_values = {}
    if dictionary_setting == "Manual":
        dictionary_q_values = _read_dictionary_table(
            ws,
            header_row=15,
            start_col=1,
            end_col=8,
            data_start_row=16,
        )
        if not dictionary_q_values:
            raise ConfigError(
                "Dictionary Setting is Manual, but no Q values were detected in the dictionary table."
            )

    return ExcelConfig(
        optimizing_method=optimizing_method,
        init_condition_setting=init_condition_setting,
        initial_design_options=initial_design_options,
        xinit_indices=xinit_indices,
        dictionary_setting=dictionary_setting,
        dictionary_q_values=dictionary_q_values,
    )



def _x_number(x_name: str) -> int:
    """
    Extract x index from strings like 'x1', 'x8'. Unknown -> large.
    """
    m = re.match(r"^\s*x\s*(\d+)\s*$", str(x_name), flags=re.IGNORECASE)
    return int(m.group(1)) if m else 10**9


def _sorted_vector_names_by_x(capture_result) -> List[str]:
    """
    capture_result.vectors keys are like 'x1_Q1_value', 'x8_Q2_value'
    We sort by the x number, increasing: x1 ... x12
    """
    items: List[Tuple[int, str]] = []
    for vec_name, vec in capture_result.vectors.items():
        items.append((_x_number(vec.x_name), vec_name))
    items.sort(key=lambda t: t[0])
    return [vn for _, vn in items]


def _total_combinations(sizes: List[int]) -> int:
    total = 1
    for s in sizes:
        total *= int(s)
    return total


def iter_index_rows_last_fastest(sizes: List[int]) -> Iterable[List[int]]:
    """
    Yield rows of indices (1-based) for each variable.
    The LAST variable changes fastest.

    Example sizes [2,3] yields:
      [1,1], [1,2], [1,3], [2,1], [2,2], [2,3]
    """
    if not sizes:
        return
    if any(s <= 0 for s in sizes):
        raise ValueError(f"All sizes must be > 0, got {sizes}")

    n = len(sizes)
    idx = [1] * n  # 1-based

    while True:
        yield idx.copy()

        # increment from last to first (last is fastest)
        for k in range(n - 1, -1, -1):
            idx[k] += 1
            if idx[k] <= sizes[k]:
                break
            idx[k] = 1
            if k == 0:
                return

def write_design_option_index_matrix(
    capture_result,
    output_xlsx_path: str,
    *,
    sheet_name: str = "DesignOptions_Index",
    create_new_workbook: bool = False,
    max_rows: int = 2_000_000,     # safety guard; set None to disable
    include_header: bool = True,
    include_order_col: bool = True,
    z_index_mode: Literal["generator", "list", "numpy"] = "generator",
    write_excel: bool = True,
) -> Dict[str, object]:
    """
    Generate and save the index matrix based on capture_result.

    Columns correspond to variables sorted by x number (x1..x12).
    Each cell is the 1-based index into that variable's vector (not the value itself).

    Returns a summary dict with:
      - vector_names_in_order
      - sizes
      - total_combinations
      - rows_written
      - output_xlsx_path
      - sheet_name

    Create the same index matrix in:
      - Python (Z_index)
      - and optionally Excel sheet

    z_index_mode:
      - "generator": return an iterator (memory-safe)
      - "list": return List[List[int]]
      - "numpy": return np.ndarray (requires total rows not huge)
    """
    vector_names = _sorted_vector_names_by_x(capture_result)
    if not vector_names:
        raise ValueError("No variables with data found in capture_result.")

    sizes = [capture_result.sizes[vn] for vn in vector_names]
    total = 1
    for s in sizes:
        total *= int(s)

    if total > max_rows:
        raise ValueError(
            f"Too many design options: {total:,} rows exceed max_rows={max_rows:,}."
        )

    # ---- Build Z_index in Python ----

    rows = list(iter_index_rows_last_fastest(sizes))

    if z_index_mode == "generator":
        Z_index = (row for row in rows)  # re-generator
    elif z_index_mode == "list":
        Z_index = rows
    elif z_index_mode == "numpy":
        Z_index = np.asarray(rows, dtype=np.int32)
    else:
        raise ValueError(f"Unknown z_index_mode: {z_index_mode}")

    if write_excel:
        if not os.path.exists(output_xlsx_path):
            raise FileNotFoundError(f"Target Excel file not found: {output_xlsx_path}")

        wb = load_workbook(output_xlsx_path)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in '{output_xlsx_path}'.")

        ws = wb[sheet_name]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)

        if include_header:
            headers = [capture_result.vectors[vn].x_name for vn in vector_names]
            if include_order_col:
                headers = ["Order"] + headers
            ws.append(headers)
            ws.freeze_panes = "A2"

        if include_order_col:
            for i, row in enumerate(rows, start=1):
                ws.append([i] + row)
        else:
            for row in rows:
                ws.append(row)

        wb.save(output_xlsx_path)

    return {
        "vector_names_in_order": vector_names,
        "sizes": sizes,
        "total_combinations": total,
        "Z_index": Z_index,
        "z_index_mode": z_index_mode,
    }


def device_distance_by_no(
    file_path: Union[str, Path],
    sheet_name: str,
    no_x: int,
    no_y: int,
    *,
    feature_cols: Sequence[str],
    weights: Optional[Union[Sequence[float], Mapping[str, float]]] = None,
    no_col: str = "Index",
    device_col: str = "Device",
    block_col: str = "BlockID",
    block_x: Optional[int] = None,
    block_y: Optional[int] = None,
) -> dict:

    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # checks
    for required in (no_col, block_col):
        if required not in df.columns:
            raise ValueError(f"Column '{required}' not found in the sheet.")

    missing = [c for c in feature_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing feature columns: {missing}")

    df[no_col] = pd.to_numeric(df[no_col], errors="coerce")
    df[block_col] = pd.to_numeric(df[block_col], errors="coerce")

    # select rows (block-aware)
    qx = (df[no_col] == no_x)
    qy = (df[no_col] == no_y)

    if block_x is not None:
        qx &= (df[block_col] == block_x)
    if block_y is not None:
        qy &= (df[block_col] == block_y)

    row_x = df.loc[qx]
    row_y = df.loc[qy]

    # ambiguity handling
    if row_x.empty:
        raise ValueError(f"Device not found for ({block_col}={block_x}, {no_col}={no_x}).")
    if row_y.empty:
        raise ValueError(f"Device not found for ({block_col}={block_y}, {no_col}={no_y}).")

    if len(row_x) > 1:
        raise ValueError(
            f"Ambiguous selection for X: {len(row_x)} rows match {no_col}={no_x}. "
            f"Provide block_x or add unique key."
        )
    if len(row_y) > 1:
        raise ValueError(
            f"Ambiguous selection for Y: {len(row_y)} rows match {no_col}={no_y}. "
            f"Provide block_y or add unique key."
        )

    row_x = row_x.iloc[0]
    row_y = row_y.iloc[0]

    x = row_x[list(feature_cols)].astype(float).to_numpy()
    y = row_y[list(feature_cols)].astype(float).to_numpy()

    # weights vector
    if weights is None:
        w = np.ones(len(feature_cols), dtype=float)
    elif isinstance(weights, Mapping):
        w = np.array([float(weights.get(c, 1.0)) for c in feature_cols], dtype=float)
    else:
        w = np.asarray(weights, dtype=float)
        if w.shape != (len(feature_cols),):
            raise ValueError(f"`weights` must match len(feature_cols)={len(feature_cols)}")

    diff = x - y
    distance = float(np.sqrt(np.sum(w * (diff ** 2))))

    return {
        "block_x": int(row_x[block_col]) if not pd.isna(row_x[block_col]) else None,
        "block_y": int(row_y[block_col]) if not pd.isna(row_y[block_col]) else None,
        "No_x": int(no_x),
        "No_y": int(no_y),
        "Device_x": row_x[device_col] if device_col in df.columns else None,
        "Device_y": row_y[device_col] if device_col in df.columns else None,
        "feature_cols": list(feature_cols),
        "weights": [float(v) for v in w.tolist()],
        "deltas": {feature_cols[i]: float(diff[i]) for i in range(len(feature_cols))},
        "distance": distance,
    }

def device_distance_by_no_df(
    df: pd.DataFrame,
    no_x: int,
    no_y: int,
    *,
    feature_cols: Sequence[str],
    weights: Optional[Union[Sequence[float], Mapping[str, float]]] = None,
    no_col: str = "Index",
    device_col: str = "Device",
    block_col: str = "BlockID",
    block_x: Optional[int] = None,
    block_y: Optional[int] = None,
) -> dict:
    """
    Compute weighted distance between two devices using a PRELOADED DataFrame.
    NO Excel I/O here.
    """

    # ---- checks ----
    for required in (no_col, block_col):
        if required not in df.columns:
            raise ValueError(f"Column '{required}' not found in DataFrame.")

    missing = [c for c in feature_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing feature columns: {missing}")

    # ---- select rows (block-aware) ----
    qx = (df[no_col] == no_x)
    qy = (df[no_col] == no_y)

    if block_x is not None:
        qx &= (df[block_col] == block_x)
    if block_y is not None:
        qy &= (df[block_col] == block_y)

    row_x = df.loc[qx]
    row_y = df.loc[qy]

    if row_x.empty:
        raise ValueError(f"Device X not found: ({block_col}={block_x}, {no_col}={no_x})")
    if row_y.empty:
        raise ValueError(f"Device Y not found: ({block_col}={block_y}, {no_col}={no_y})")

    if len(row_x) > 1:
        raise ValueError(
            f"Ambiguous X selection ({len(row_x)} rows). Provide block_x."
        )
    if len(row_y) > 1:
        raise ValueError(
            f"Ambiguous Y selection ({len(row_y)} rows). Provide block_y."
        )

    row_x = row_x.iloc[0]
    row_y = row_y.iloc[0]

    # ---- feature vectors ----
    x = row_x[list(feature_cols)].astype(float).to_numpy()
    y = row_y[list(feature_cols)].astype(float).to_numpy()

    # ---- weights ----
    if weights is None:
        w = np.ones(len(feature_cols), dtype=float)
    elif isinstance(weights, Mapping):
        w = np.array([float(weights.get(c, 1.0)) for c in feature_cols], dtype=float)
    else:
        w = np.asarray(weights, dtype=float)
        if w.shape != (len(feature_cols),):
            raise ValueError("weights length must match feature_cols")

    diff = x - y
    distance = float(np.sqrt(np.sum(w * (diff ** 2))))

    return {
        "block_x": int(row_x[block_col]),
        "block_y": int(row_y[block_col]),
        "No_x": int(no_x),
        "No_y": int(no_y),
        "Device_x": row_x[device_col] if device_col in df.columns else None,
        "Device_y": row_y[device_col] if device_col in df.columns else None,
        "distance": distance,
    }


def build_index_to_block_map(
    df_feat: pd.DataFrame,
    *,
    no_col: str = "Index",
    block_col: str = "BlockID",
) -> dict:
    """
    Returns mapping: device_index(int) -> block_id(int), assuming each (Index, BlockID) pair exists.
    If the same Index appears in multiple blocks, this returns a dict of lists, but here we assume
    your Z_Index chooses Index only, so you must decide block elsewhere.
    """
    tmp = df_feat[[no_col, block_col]].dropna().copy()
    tmp[no_col] = pd.to_numeric(tmp[no_col], errors="coerce")
    tmp[block_col] = pd.to_numeric(tmp[block_col], errors="coerce")
    tmp = tmp.dropna().astype({no_col: int, block_col: int})

    # Index may repeat across blocks -> collect possible blocks
    m = {}
    for idx, blk in zip(tmp[no_col], tmp[block_col]):
        m.setdefault(int(idx), set()).add(int(blk))

    # convert sets to sorted lists for determinism
    return {k: sorted(list(v)) for k, v in m.items()}


def build_z_encoding_matrix_and_write_excel(
    *,
    start_col_index: int,                 # e.g. 5 means start writing at column E
    Z_Index: np.ndarray,                  # shape (num_vars, N_options) OR (N_options, num_vars)
    file_path: Union[str, Path],          # excel that contains features & BlockID
    features_sheet: str,                  # e.g. "CAPACITANCE"
    feature_cols: Sequence[str],          # e.g. FEATURES_Q1
    dic_list_no: Sequence[int],           # e.g. [1,2,5,6,9,19]
    weights: Optional[Union[Sequence[float], Mapping[str, float]]] = None,  # theta
    device_var_col: int = 0,
    no_col: str = "Index",
    device_col: str = "Device",
    block_col: str = "BlockID",
    # dictionary block control:
    dict_block: Optional[int] = None,     # if None, use same block as device (recommended)
    # destination:
    dest_xlsx_path: Union[str, Path],
    dest_sheet: str,
    include_header: bool = True,
    row_offset: int = 1,                  # 1-based excel row where to start writing header/values
) -> np.ndarray:
    """
    Build Z_encoding matrix (N_options, K_dict) where each entry is:
        distance(device_index_in_Z, dictionary_device_index)
    using device_distance_by_no(...) with BlockID awareness.

    Also writes the matrix into an existing Excel sheet starting at start_col_index.
    Does NOT delete sheets or recreate workbook.

    Returns:
        Z_encoding: np.ndarray shape (N_options, K_dict)
    """
    df_feat = pd.read_excel(file_path, sheet_name=features_sheet)
    df_feat[no_col] = pd.to_numeric(df_feat[no_col], errors="coerce")
    df_feat[block_col] = pd.to_numeric(df_feat[block_col], errors="coerce")

    file_path = str(file_path)
    dest_xlsx_path = str(dest_xlsx_path)

    # ---- Normalize Z_Index shape to (N_options, num_vars) ----
    Z = np.asarray(Z_Index)
    if Z.ndim != 2:
        raise ValueError("Z_Index must be 2D array.")

    # Your previous usage suggests Z_Index was (num_vars, N_options)
    # We detect orientation by assuming N_options is the larger dimension typically.
    if Z.shape[0] < Z.shape[1]:
        # likely (num_vars, N_options) -> transpose to (N_options, num_vars)
        Z_opt = Z.T
    else:
        Z_opt = Z

    N_options = Z_opt.shape[0]
    if Z_opt.shape[1] < 1:
        raise ValueError("Z_Index has no columns/variables.")

    # By your spec: first column in Z_Index is device index
    # dev_indices = Z_opt[:, 0].astype(int)
    if not (0 <= device_var_col < Z_opt.shape[1]):
        raise ValueError(
            f"device_var_col={device_var_col} out of range for Z_Index with {Z_opt.shape[1]} columns."
        )

    dev_indices = Z_opt[:, device_var_col].astype(int)
    # ---- Load feature sheet (contains Index + BlockID + normalized feature columns) ----
    # df_feat = pd.read_excel(file_path, sheet_name=features_sheet)

    # checks
    for required in (no_col, block_col):
        if required not in df_feat.columns:
            raise ValueError(f"Column '{required}' not found in features sheet '{features_sheet}'.")

    missing = [c for c in feature_cols if c not in df_feat.columns]
    if missing:
        raise ValueError(f"Missing feature columns in '{features_sheet}': {missing}")

    # Build Index -> possible blocks mapping
    idx_to_blocks = build_index_to_block_map(df_feat, no_col=no_col, block_col=block_col)

    # ---- Compute Z_encoding (one-by-one distance calls) ----
    K = len(dic_list_no)
    Z_encoding = np.empty((N_options, K), dtype=float)

    for r in range(N_options):
        no_x = int(dev_indices[r])

        # Determine block for this device index
        blocks = idx_to_blocks.get(no_x, [])
        if not blocks:
            raise ValueError(f"Device Index={no_x} not found in '{features_sheet}'.")

        if len(blocks) > 1:
            # Ambiguous because Index repeats across blocks.
            # Here we choose the first by default, but better: pass block info alongside Z_Index.
            # You can also change this to raise an error instead.
            block_x = blocks[0]
        else:
            block_x = blocks[0]

        for j, no_y in enumerate(dic_list_no):
            no_y = int(no_y)

            # Dictionary block logic:
            # - If dict_block is specified, use it
            # - else use same block as the device (recommended, "same block consideration")
            block_y = dict_block if dict_block is not None else block_x

            res = device_distance_by_no_df(
                df_feat,
                no_x=no_x,
                no_y=no_y,
                block_x=block_x,
                block_y=block_y,
                feature_cols=feature_cols,
                weights=weights,
                no_col=no_col,
                device_col=device_col,
                block_col=block_col,
            )
            Z_encoding[r, j] = float(res["distance"])

    # ---- Write to destination Excel (existing workbook/sheet) ----
    wb = load_workbook(dest_xlsx_path)
    if dest_sheet not in wb.sheetnames:
        raise ValueError(f"Destination sheet '{dest_sheet}' not found in '{dest_xlsx_path}'.")

    ws = wb[dest_sheet]

    start_col = int(start_col_index)
    if start_col < 1:
        raise ValueError("start_col_index must be >= 1 (Excel columns are 1-based).")

    # Write header row (optional)
    write_row = int(row_offset)
    if include_header:
        for j, no_y in enumerate(dic_list_no):
            c = start_col + j
            ws.cell(write_row, c).value = f"dist_to_{no_y}"
        write_row += 1  # data begins next row

    # Write data rows
    for r in range(N_options):
        for j in range(K):
            ws.cell(write_row + r, start_col + j).value = float(Z_encoding[r, j])

    wb.save(dest_xlsx_path)

    return Z_encoding

#
# # Assuming you still have this helper from your previous context,
# # or you can inline the logic if it's simple.
# # We need it to handle the BlockID mapping.
# def build_index_to_block_map(df, no_col, block_col):
#     mapping = {}
#     for idx, blk in zip(df[no_col], df[block_col]):
#         if pd.isna(idx) or pd.isna(blk): continue
#         idx = int(idx)
#         blk = int(blk)
#         if idx not in mapping:
#             mapping[idx] = []
#         if blk not in mapping[idx]:
#             mapping[idx].append(blk)
#     return mapping


def build_feature_matrix_and_write_excel(
        *,
        start_col_index: int,  # e.g. 5 means start writing at column E
        Z_Index: np.ndarray,  # shape (num_vars, N_options) OR (N_options, num_vars)
        file_path: Union[str, Path],  # excel that contains features & BlockID
        features_sheet: str,  # e.g. "CAPACITANCE"
        feature_cols: Sequence[str],  # e.g. ["RdsOn", "Qg", "Coss"] -> These will be the columns
        device_var_col: int = 0,
        no_col: str = "Index",
        block_col: str = "BlockID",
        # destination:
        dest_xlsx_path: Union[str, Path],
        dest_sheet: str,
        include_header: bool = True,
        row_offset: int = 1,  # 1-based excel row where to start writing header/values
) -> np.ndarray:
    """
    Builds a matrix (N_options, N_features) by extracting raw feature values
    for the devices specified in Z_Index.

    Writes the matrix into an existing Excel sheet starting at start_col_index.
    """
    # 1. Load Data
    df_feat = pd.read_excel(file_path, sheet_name=features_sheet)
    #Remove all non-numeric data in "no_col" and "block_col" by transfer them to NaN
    df_feat[no_col] = pd.to_numeric(df_feat[no_col], errors="coerce")
    df_feat[block_col] = pd.to_numeric(df_feat[block_col], errors="coerce")

    # Ensure feature columns are numeric (optional, but good for safety)
    for col in feature_cols:
        df_feat[col] = pd.to_numeric(df_feat[col], errors="coerce")

    # file_path = str(file_path)
    dest_xlsx_path = str(dest_xlsx_path)

    # 2. Normalize Z_Index shape to (N_options, num_vars)
    Z = np.asarray(Z_Index)
    if Z.ndim != 2:
        raise ValueError("Z_Index must be 2D array.")

    #If more columns than rows, then need transposing Z.
    if Z.shape[0] < Z.shape[1]:
        Z_opt = Z.T
    else:
        Z_opt = Z

    N_options = Z_opt.shape[0]
    if Z_opt.shape[1] < 1:
        raise ValueError("Z_Index has no columns/variables.")

    if not (0 <= device_var_col < Z_opt.shape[1]):
        raise ValueError(f"device_var_col={device_var_col} out of range.")

    # Device index 1D array for a specific column "device_var_col" in Zopt
    dev_indices = Z_opt[:, device_var_col].astype(int)

    # 3. Validation
    for required in (no_col, block_col):
        if required not in df_feat.columns:
            raise ValueError(f"Column '{required}' not found in features sheet.")

    missing = [c for c in feature_cols if c not in df_feat.columns]
    if missing:
        raise ValueError(f"Missing feature columns in '{features_sheet}': {missing}")

    # Build Index -> possible blocks mapping (to handle duplicates)
    idx_to_blocks = build_index_to_block_map(df_feat, no_col=no_col, block_col=block_col)
    # print("idx_to_blocks:",idx_to_blocks)
    # 4. Extract Features (The Core Logic Change)
    num_features = len(feature_cols)
    Z_output = np.empty((N_options, num_features), dtype=float)

    # Pre-indexing for speed: Create a multi-index lookup if possible,
    # but strictly following your BlockID logic is safer line-by-line.
    #
    for r in range(N_options):
        no_x = int(dev_indices[r])

        # Determine block for this device index
        blocks = idx_to_blocks.get(no_x, [])
        # print("blocks:",blocks)
        if not blocks:
            raise ValueError(f"Device Index={no_x} not found in '{features_sheet}'.")

        # Resolve ambiguity (Default to first block found, same as previous logic)
        block_x = blocks[device_var_col]
        # print("block_x:", block_x)
        # Find the specific row in the DataFrame
        # We look for the exact match of Index AND BlockID
        mask = (df_feat[no_col] == no_x) & (df_feat[block_col] == block_x)
        target_row = df_feat.loc[mask]

        if target_row.empty:
            raise ValueError(f"Device {no_x} in Block {block_x} not found in data.")

        # Extract values for the requested feature columns
        # .iloc[0] ensures we take the values as a Series/array, not a DataFrame
        values = target_row[feature_cols].iloc[0].values

        # Store in matrix
        Z_output[r, :] = values
    #
    # 5. Write to destination Excel
    wb = load_workbook(dest_xlsx_path)
    if dest_sheet not in wb.sheetnames:
        raise ValueError(f"Destination sheet '{dest_sheet}' not found.")

    ws = wb[dest_sheet]
    start_col = int(start_col_index)
    write_row = int(row_offset)

    # Write Header (Feature Names)
    if include_header:
        for j, col_name in enumerate(feature_cols):
            ws.cell(write_row, start_col + j).value = col_name
        write_row += 1

    # Write Data Rows
    for r in range(N_options):
        for j in range(num_features):
            ws.cell(write_row + r, start_col + j).value = float(Z_output[r, j])

    wb.save(dest_xlsx_path)

    return Z_output


def build_normalized_input_block_and_write_excel(
    *,
    start_col_index: int,
    Z_Index: np.ndarray,
    vec,                         # VariableVector
    device_var_col: int,         # column in Z_Index to use
    dest_xlsx_path: str,
    dest_sheet: str,
    include_header: bool = True,
    row_offset: int = 1,
    write_excel: bool = True,    # NEW
) -> np.ndarray:
    """
    Build a normalized input column for parameter / magnetic variables.

    - Always RETURNS Z_norm: shape (N_options, 1)
    - Optionally WRITES to Excel if write_excel=True
    """

    import numpy as np
    from openpyxl import load_workbook

    # ---- normalize Z_Index shape ----
    Z = np.asarray(Z_Index)
    if Z.ndim != 2:
        raise ValueError("Z_Index must be 2D array.")
    if Z.shape[0] < Z.shape[1]:
        Z = Z.T  # (N_options, num_vars)

    if not (0 <= device_var_col < Z.shape[1]):
        raise ValueError(f"device_var_col={device_var_col} out of range for Z_Index with {Z.shape[1]} cols.")

    idx_col = Z[:, device_var_col].astype(int)  # 1-based indices

    values = np.asarray(vec.values, dtype=float)
    if values.size == 0:
        raise ValueError(f"No values found for {vec.x_name}")

    # ---- index lookup (1-based) ----
    if np.any(idx_col < 1) or np.any(idx_col > values.size):
        bad = idx_col[(idx_col < 1) | (idx_col > values.size)]
        raise ValueError(
            f"Z_Index contains invalid indices for {vec.x_name}. "
            f"Valid range: 1..{values.size}. Bad samples: {bad[:10].tolist()}"
        )

    raw_vals = values[idx_col - 1]

    # ---- normalization ----
    mean_val = float(np.mean(values))
    if mean_val == 0:
        raise ValueError(f"Mean is zero for {vec.x_name}, cannot normalize")

    Z_norm = (raw_vals / mean_val).reshape(-1, 1)  # (N, 1)

    # ---- write to Excel (optional) ----
    if write_excel:
        wb = load_workbook(dest_xlsx_path)
        if dest_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{dest_sheet}' not found.")

        ws = wb[dest_sheet]
        r0 = int(row_offset)

        if include_header:
            ws.cell(r0, start_col_index).value = f"{vec.x_name}_norm"
            r0 += 1

        for i, v in enumerate(Z_norm[:, 0]):
            ws.cell(r0 + i, start_col_index).value = float(v)

        wb.save(dest_xlsx_path)

    return Z_norm
