import argparse
import random
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.paths import RESULT_PATH


# -----------------------------
# Helpers
# -----------------------------
def find_last_data_row(ws: Worksheet, index_col: int = 1, header_row: int = 1) -> int:
    """Last non-empty row in index column (default: Column A)."""
    r = ws.max_row
    while r > header_row and ws.cell(r, index_col).value in (None, ""):
        r -= 1
    return r


def detect_block_size(ws: Worksheet, device_col: int = 2, start_row: int = 2, max_scan: int = 50000) -> int:
    """
    Detect how many consecutive rows belong to the first device before device changes.
    This should be 96 in your case (8*12).
    """
    first = ws.cell(start_row, device_col).value
    if first in (None, ""):
        raise ValueError("Cannot detect block size: first device cell is empty.")

    r = start_row
    for _ in range(max_scan):
        r += 1
        v = ws.cell(r, device_col).value
        if v != first:
            return r - start_row

    raise ValueError("Cannot detect block size within scan limit. Check sheet structure.")


def extract_devices(ws: Worksheet, start_row: int, last_row: int, device_col: int, block_size: int) -> List[str]:
    """Read device name at the top of each block."""
    devices = []
    r = start_row
    while r <= last_row:
        dev = ws.cell(r, device_col).value
        if dev in (None, ""):
            raise ValueError(f"Empty device at row {r}; cannot build device list.")
        devices.append(str(dev).strip())
        r += block_size
    return devices


def write_devices_by_blocks(ws: Worksheet, devices_shuffled: List[str], start_row: int, last_row: int,
                           device_col: int, block_size: int) -> None:
    """Overwrite Column B device values in blocks, leaving all other columns unchanged."""
    total_rows = last_row - start_row + 1
    if total_rows % block_size != 0:
        raise ValueError(f"Data rows ({total_rows}) not divisible by block_size ({block_size}).")

    n_blocks = total_rows // block_size
    if n_blocks != len(devices_shuffled):
        raise ValueError(f"Block count mismatch: detected {n_blocks} blocks but got {len(devices_shuffled)} devices.")

    r = start_row
    for dev in devices_shuffled:
        for rr in range(r, r + block_size):
            ws.cell(rr, device_col).value = dev
        r += block_size


def find_header_row(ws: Worksheet, header_text: str, col: int, scan_rows: int = 50) -> Optional[int]:
    """Find header row by matching text in a given column (case-insensitive)."""
    target = header_text.strip().lower()
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and v.strip().lower() == target:
            return r
    return None


def update_yellow_part_list(ws: Worksheet, devices_shuffled: List[str],
                           no_col: int = 10, part_col: int = 11) -> None:
    """
    Update the right-side list in J:K if it exists:
    J header: "No"
    K header: "Part Number"
    """
    hdr_row = find_header_row(ws, "No", col=no_col, scan_rows=80)
    if hdr_row is None:
        # Sheet might not have that list area; silently skip.
        return

    start = hdr_row + 1
    for i, dev in enumerate(devices_shuffled, start=1):
        ws.cell(start + i - 1, no_col).value = i
        ws.cell(start + i - 1, part_col).value = dev


def clear_columns(ws: Worksheet, start_row: int, last_row: int, cols_to_clear: List[int]) -> None:
    """Clear values only (keep formatting)."""
    for r in range(start_row, last_row + 1):
        for c in cols_to_clear:
            ws.cell(r, c).value = None


# -----------------------------
# Main
# -----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--in", dest="in_path",
        default=RESULT_PATH,
        help="Input .xlsx path"
    )
    ap.add_argument(
        "--out", dest="out_path",
        default=RESULT_PATH,
        help="Output .xlsx path"
    )
    ap.add_argument("--template", default="Test Case 01", help="Template sheet name")
    ap.add_argument("--n_new", type=int, default=8, help="Number of new sheets to create (default 8 => 02..09)")
    ap.add_argument("--seed", type=int, default=123, help="Random seed (optional)")
    ap.add_argument("--block_size", type=int, default=0, help="Rows per device block. Set 96 to force. 0 => auto-detect.")
    ap.add_argument("--clear_results", action="store_true", help="Clear result columns (Efficiency/status/error/timestamp) in new sheets")

    args = ap.parse_args()

    if args.seed is not None:
        random.seed(args.seed)

    # Use args everywhere
    in_path = Path(args.in_path)
    out_path = Path(args.out_path)
    template_name = "Test Case 01"

    wb = load_workbook(in_path)
    if args.template not in wb.sheetnames:
        raise ValueError(f"Template sheet '{args.template}' not found. Sheets: {wb.sheetnames}")
    tpl = wb[template_name]

    header_row = 1
    data_start_row = 2
    last_row = find_last_data_row(tpl, index_col=1, header_row=header_row)

    # Your structure: Column A = index, Column B = device, Column C = Freq_Hz, Column D = Turn
    device_col = 2

    block_size = args.block_size if args.block_size > 0 else detect_block_size(
        tpl, device_col=device_col, start_row=data_start_row
    )

    devices = extract_devices(tpl, data_start_row, last_row, device_col=device_col, block_size=block_size)

    # Create new test cases
    for k in range(2, 2 + args.n_new):
        new_name = f"Test Case {k:02d}"

        # If exists, delete and recreate
        if new_name in wb.sheetnames:
            del wb[new_name]

        ws_new = wb.copy_worksheet(tpl)
        ws_new.title = new_name

        # Shuffle device order, then write device column by blocks
        shuffled = devices[:]
        random.shuffle(shuffled)
        write_devices_by_blocks(ws_new, shuffled, data_start_row, last_row, device_col=device_col, block_size=block_size)

        # Update right-side list (J:K) if present
        update_yellow_part_list(ws_new, shuffled, no_col=10, part_col=11)

        # Optional: clear result columns (E=Efficiency, F=status, G=error, H=timestamp in your screenshot)
        if args.clear_results:
            clear_columns(ws_new, data_start_row, last_row, cols_to_clear=[5, 6, 7, 8])

    wb.save(out_path)
    print(f"Done. Saved output to: {out_path}")
    print(f"Detected/used block_size = {block_size}, devices = {len(devices)}")


if __name__ == "__main__":
    main()
