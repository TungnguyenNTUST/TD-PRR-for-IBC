"""
Dictionary-based input matrix builder for BO (Paper 01 – DIC variant).

Reads UserProvidedDataFile_DIC.xlsx, computes the dictionary-distance
encoding for each design option, and writes the result back to the
Mid_Input_encoding sheet in that same file (+ .npy cache).

Does NOT touch UserProvidedDataFile.xlsx (raw-features variant).

Encoding structure (one row per design option, N rows total):
  [dist(MOS_r, dic_MOS_1), ..., dist(MOS_r, dic_MOS_K1),   <- K1 cols (Q1)
   dist(Core_r, dic_Core_1), ..., dist(Core_r, dic_Core_K2), <- K2 cols (Q2)
   dist(Dio_r, dic_Dio_1),  ..., dist(Dio_r, dic_Dio_K3),   <- K3 cols (Q3)
   freq_norm_r,                                               <- 1 col  (x4)
   ind_norm_r]                                                <- 1 col  (x5)

Distance formula (weighted Euclidean on normalized features):
  d(a, b) = sqrt( sum_i( w_i * (a_i - b_i)^2 ) )
"""
from __future__ import annotations

import os
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from core.paths import FILE_PATH, FILE_PATH_DIC

# ── Sheet names (same as original file) ───────────────────────────────────────
SHEET_NORM    = "Mid_Normalization"
SHEET_INDEX   = "Mid_Input_Indexing"
SHEET_OPT     = "Optimization"
SHEET_DEVICE  = "Device"
SHEET_ENC_OUT = "Mid_Input_encoding"

# ── Cache file for Input matrix (.npy) ────────────────────────────────────────
_DIC_DIR = os.path.dirname(FILE_PATH_DIC)
CACHE_INPUT_NPY = os.path.join(_DIC_DIR, "Mid_Input_encoding_DIC.npy")
CACHE_INDEX_NPY = os.path.join(_DIC_DIR, "Mid_Input_Indexing_DIC.npy")

# ── Device-block mapping ───────────────────────────────────────────────────────
# Must match the BlockID values in Mid_Normalization and column order in
# Mid_Input_Indexing:  Order | x1 | x2 | x3 | x4 | x5
#                             Q1   Q2   Q3   freq  ind
BLOCK_TO_X_COL = {1: 1, 2: 2, 3: 3}   # BlockID -> Z_index column (1-based col index)
X4_COL = 4   # switching frequency column in Z_index
X5_COL = 5   # inductance column in Z_index


# ── 1. Parse Device sheet: features and weights per Q ─────────────────────────

def parse_device_sheet(excel_path: str) -> Dict[str, Dict]:
    """
    Read the Device sheet and return:
        {"Q1": {"features": [normalized_col_names], "weights": [floats]},
         "Q2": {...}, "Q3": {...}}

    The sheet layout:
        Row 0: header (ignored)
        Row 1: NaN | Component | Q1 | Q2 | Q3 | ...
        Row 2+: W_i | feature_name | w_Q1 | w_Q2 | w_Q3 | ...
    """
    df = pd.read_excel(excel_path, sheet_name=SHEET_DEVICE, header=None)

    # Locate the header row (contains "Q1")
    header_row = None
    for i, row in df.iterrows():
        if "Q1" in row.values:
            header_row = i
            break
    if header_row is None:
        raise ValueError("Cannot find 'Q1' header in Device sheet.")

    q_cols: Dict[str, int] = {}
    for col_idx, val in enumerate(df.iloc[header_row]):
        s = str(val).strip()
        if s.startswith("Q") and s[1:].isdigit():
            q_cols[s] = col_idx   # e.g. {"Q1": 2, "Q2": 3, "Q3": 4}

    feature_name_col = 1
    data_rows = df.iloc[header_row + 1 :]

    result: Dict[str, Dict] = {q: {"features": [], "weights": []} for q in q_cols}

    for _, row in data_rows.iterrows():
        feat_name = str(row.iloc[feature_name_col]).strip()
        if feat_name in ("nan", "", "None"):
            continue
        norm_col = f"{feat_name}_normalized"
        for q, col_idx in q_cols.items():
            w = row.iloc[col_idx]
            if pd.notna(w) and float(w) != 0:
                result[q]["features"].append(norm_col)
                result[q]["weights"].append(float(w))

    # Keep only Qs that have at least one feature
    return {q: v for q, v in result.items() if v["features"]}


# ── 2. Parse Optimization sheet: dictionary indices per Q ─────────────────────

def parse_dictionary_config(excel_path: str) -> Dict[str, List[int]]:
    """
    Read the Optimization sheet and return dictionary indices per Q.
        {"Q1": [1, 3, 5, ...], "Q2": [...], "Q3": [...]}

    Expected layout:
        ... (rows above)
        Row N:   "Dictionary Setting" | "Manual"
        Row N+1: "Q1" | "Q2" | "Q3" | ...   <- column headers for dictionary table
        Row N+2+: int values (NaN = no entry)
    """
    df = pd.read_excel(excel_path, sheet_name=SHEET_OPT, header=None)

    # Find the "Dictionary Setting" row
    dict_header_row = None
    for i, row in df.iterrows():
        for val in row.values:
            if str(val).strip() == "Dictionary Setting":
                dict_header_row = i
                break
        if dict_header_row is not None:
            break

    if dict_header_row is None:
        raise ValueError("Cannot find 'Dictionary Setting' row in Optimization sheet.")

    # Next row has Q column headers
    q_header_row = dict_header_row + 1
    q_cols: Dict[str, int] = {}
    for col_idx, val in enumerate(df.iloc[q_header_row]):
        s = str(val).strip()
        if s.startswith("Q") and s[1:].isdigit():
            q_cols[s] = col_idx

    # Rows below: dictionary index values
    dic_config: Dict[str, List[int]] = {q: [] for q in q_cols}
    for _, row in df.iloc[q_header_row + 1 :].iterrows():
        for q, col_idx in q_cols.items():
            val = row.iloc[col_idx]
            if pd.notna(val):
                dic_config[q].append(int(val))

    return {q: v for q, v in dic_config.items() if v}


# ── 3. Load Mid_Normalization features ────────────────────────────────────────

def load_normalization(excel_path: str) -> pd.DataFrame:
    """Return full Mid_Normalization dataframe (Index, Device, BlockID, *_normalized)."""
    df = pd.read_excel(excel_path, sheet_name=SHEET_NORM)
    required = {"Index", "BlockID"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Mid_Normalization missing columns: {missing}")
    return df


# ── 4. Load Z_index (Mid_Input_Indexing) ─────────────────────────────────────

def load_z_index(excel_path: str) -> np.ndarray:
    """Return Z_index as (N, n_vars) int array. Columns: [Order, x1, x2, x3, x4, x5]."""
    df = pd.read_excel(excel_path, sheet_name=SHEET_INDEX, header=0)
    return df.to_numpy(dtype=int)   # (N, 6)


# ── 5. Precompute device-to-dictionary distance table ─────────────────────────

def _weighted_euclidean(a: np.ndarray, b: np.ndarray, w: np.ndarray) -> float:
    diff = a - b
    return float(np.sqrt(np.sum(w * diff ** 2)))


def build_device_dist_table(
    df_norm: pd.DataFrame,
    block_id: int,
    feature_cols: List[str],
    weights: List[float],
    dic_indices: List[int],
) -> np.ndarray:
    """
    Return (n_devices, K) distance matrix where:
        n_devices = number of devices in this block
        K = len(dic_indices)

    Row i = distances from device (Index=i+1) to each dictionary device.
    All devices AND dictionary entries are from the same block (same-block constraint).
    """
    df_block = df_norm[df_norm["BlockID"] == block_id].copy()
    df_block = df_block.sort_values("Index").reset_index(drop=True)

    missing_feats = [f for f in feature_cols if f not in df_block.columns]
    if missing_feats:
        raise ValueError(f"Block {block_id}: missing feature columns {missing_feats}")

    n_devices = len(df_block)
    K = len(dic_indices)
    w = np.array(weights, dtype=float)

    # Feature matrix for all devices in this block: (n_devices, n_feats)
    F_all = df_block[feature_cols].to_numpy(dtype=float)

    # Feature matrix for dictionary devices (same block)
    dic_rows = []
    for dic_idx in dic_indices:
        rows = df_block[df_block["Index"] == dic_idx]
        if rows.empty:
            raise ValueError(
                f"Dictionary index {dic_idx} not found in Block {block_id} "
                f"(available: {df_block['Index'].tolist()})"
            )
        dic_rows.append(rows[feature_cols].to_numpy(dtype=float)[0])
    F_dic = np.vstack(dic_rows)  # (K, n_feats)

    # Vectorized weighted Euclidean: broadcast (n_devices, 1, n_feats) - (1, K, n_feats)
    diff = F_all[:, np.newaxis, :] - F_dic[np.newaxis, :, :]   # (n_devices, K, n_feats)
    dist_table = np.sqrt(np.sum(w[np.newaxis, np.newaxis, :] * diff ** 2, axis=2))  # (n_devices, K)
    return dist_table   # index device (0-based) -> row


# ── 6. Build full encoding matrix ─────────────────────────────────────────────

def build_encoding_matrix(
    z_index: np.ndarray,
    df_norm: pd.DataFrame,
    device_features: Dict[str, Dict],
    dic_config: Dict[str, List[int]],
) -> Tuple[np.ndarray, List[str]]:
    """
    Build the (N, D) encoding matrix and the list of D column headers.

    Column order:
        Q1 distances | Q2 distances | Q3 distances | freq_norm | ind_norm

    z_index columns: [Order, x1, x2, x3, x4, x5]
    """
    N = z_index.shape[0]
    blocks_ordered = sorted(BLOCK_TO_X_COL.keys())   # [1, 2, 3]

    # Map Q name -> block id: Q1->1, Q2->2, Q3->3
    q_to_block = {f"Q{b}": b for b in blocks_ordered}

    enc_blocks: list[np.ndarray] = []
    headers: list[str] = []

    for block_id in blocks_ordered:
        q_name = f"Q{block_id}"
        if q_name not in dic_config:
            raise ValueError(
                f"Dictionary not configured for {q_name}. "
                f"Please fill in the {q_name} column in the Optimization sheet."
            )
        if q_name not in device_features:
            raise ValueError(
                f"No features selected for {q_name} in the Device sheet."
            )

        dic_indices = dic_config[q_name]
        feat_cols   = device_features[q_name]["features"]
        weights     = device_features[q_name]["weights"]
        x_col       = BLOCK_TO_X_COL[block_id]   # column in z_index (1-based)

        print(f"  [{q_name}] block={block_id}, "
              f"features={len(feat_cols)}, dict_size={len(dic_indices)}")

        # Precompute (n_devices_in_block, K) distance table
        dist_table = build_device_dist_table(df_norm, block_id, feat_cols, weights, dic_indices)

        # x_col is 1-based into z_index columns [Order, x1, x2, x3, x4, x5]
        device_indices = z_index[:, x_col]   # 1-based device index per option

        # Map each design option to its row in dist_table (0-based)
        enc_block = dist_table[device_indices - 1, :]   # (N, K)
        enc_blocks.append(enc_block)

        for dic_idx in dic_indices:
            headers.append(f"{q_name}_dist_to_{dic_idx}")

    # Normalize x4 (freq) and x5 (ind) – both range 1..max
    x4 = z_index[:, X4_COL].astype(float)
    x5 = z_index[:, X5_COL].astype(float)
    x4_norm = (x4 - x4.min()) / (x4.max() - x4.min())
    x5_norm = (x5 - x5.min()) / (x5.max() - x5.min())

    enc_blocks.append(x4_norm[:, np.newaxis])
    enc_blocks.append(x5_norm[:, np.newaxis])
    headers.append("x4_norm")
    headers.append("x5_norm")

    encoding = np.hstack(enc_blocks)   # (N, D)
    return encoding, headers


# ── 7. Copy reference sheets from original file to DIC file ──────────────────

def copy_sheet_to_dic(src_path: str, dst_path: str, sheet_name: str) -> None:
    """
    Overwrite sheet_name in dst_path with the contents of the same sheet
    from src_path.  Both workbooks must already have the sheet.
    """
    df = pd.read_excel(src_path, sheet_name=sheet_name)

    wb_dst = load_workbook(dst_path)
    if sheet_name not in wb_dst.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {dst_path}.")
    ws = wb_dst[sheet_name]

    # Clear
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # Header
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(1, col_idx).value = col_name

    # Data
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row_data, start=1):
            ws.cell(r_idx, c_idx).value = None if pd.isna(val) else val

    wb_dst.save(dst_path)
    print(f"[COPIED] '{sheet_name}' ({len(df)} rows x {len(df.columns)} cols) -> DIC file")


# ── 8. Write encoding to Excel + save .npy ───────────────────────────────────

def write_encoding_excel(excel_path: str, encoding: np.ndarray, headers: List[str]) -> None:
    """Overwrite Mid_Input_encoding sheet with the new dictionary-encoding matrix."""
    wb = load_workbook(excel_path)
    if SHEET_ENC_OUT not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_ENC_OUT}' not found in {excel_path}.")
    ws = wb[SHEET_ENC_OUT]

    # Clear existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # Write header row
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(1, col_idx).value = h

    # Write data rows
    N, D = encoding.shape
    for r in range(N):
        for c in range(D):
            ws.cell(r + 2, c + 1).value = float(encoding[r, c])

    wb.save(excel_path)
    print(f"[SAVED] '{SHEET_ENC_OUT}' written to {excel_path}  ({N} rows × {D} cols)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    excel_path = FILE_PATH_DIC
    print(f"[INFO] DIC file: {excel_path}")

    print("[STEP 1] Parsing Device sheet (features & weights per Q)...")
    device_features = parse_device_sheet(excel_path)
    for q, v in device_features.items():
        print(f"  {q}: {v['features']}")

    print("[STEP 2] Parsing dictionary config from Optimization sheet...")
    dic_config = parse_dictionary_config(excel_path)
    for q, indices in dic_config.items():
        print(f"  {q}: {len(indices)} entries -> {indices}")

    # Mid_Normalization and Mid_Input_Indexing: read from original (always
    # populated), then mirror into the DIC file for reference / cross-checking.
    print("[STEP 3] Syncing Mid_Normalization from original -> DIC file...")
    copy_sheet_to_dic(FILE_PATH, excel_path, SHEET_NORM)
    df_norm = load_normalization(FILE_PATH)
    print(f"  {len(df_norm)} devices, blocks: {sorted(df_norm['BlockID'].unique().tolist())}")

    print("[STEP 4] Syncing Mid_Input_Indexing from original -> DIC file...")
    copy_sheet_to_dic(FILE_PATH, excel_path, SHEET_INDEX)
    z_index = load_z_index(FILE_PATH)
    print(f"  Z_index shape: {z_index.shape}  ({z_index.shape[0]} design options)")

    # Save Z_index .npy cache
    np.save(CACHE_INDEX_NPY, z_index)
    print(f"  [CACHE] Z_index saved -> {CACHE_INDEX_NPY}")

    print("[STEP 5] Building encoding matrix...")
    encoding, headers = build_encoding_matrix(z_index, df_norm, device_features, dic_config)
    print(f"  Encoding shape: {encoding.shape}  (D = {encoding.shape[1]} features)")
    print(f"  Columns: {headers}")

    # Save Input matrix .npy cache  (transposed to (D, N) for DiscreteBO)
    Input = encoding.T    # (D, N)
    np.save(CACHE_INPUT_NPY, Input)
    print(f"  [CACHE] Input matrix (D×N) saved -> {CACHE_INPUT_NPY}")

    print("[STEP 6] Writing encoding to Excel...")
    write_encoding_excel(excel_path, encoding, headers)

    print("\n[DONE] Dictionary-encoding input matrix complete.")
    print(f"  Columns (D={encoding.shape[1]}): {headers}")


if __name__ == "__main__":
    main()
